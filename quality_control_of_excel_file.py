import sys

import pandas

FILENAME = sys.argv[1] #"20220121 Overview CG plates and compounds  _consolidated RTG.xlsx"

dfs_per_sheetname = pandas.read_excel(FILENAME, sheet_name=None)

assert "experiments" in dfs_per_sheetname
df = dfs_per_sheetname["experiments"]
assert "experiment ID" in df.columns
assert "compound map see corresponding excel table" in df.columns
assert df["experiment ID"].is_unique
## non-limited list


assert "imaging campaigns" in dfs_per_sheetname
df = dfs_per_sheetname["imaging campaigns"]
assert "imaging campaign ID"                             in df.columns
assert "experiment ID"                                   in df.columns
assert "timepoint in hours"                              in df.columns
assert "raw data available in zip file"                  in df.columns
assert "processed images available in folder"            in df.columns
assert "cq1 analysis available in folder"                in df.columns
assert "incucyte analyzed data available in csv file"    in df.columns
##
assert df["imaging campaign ID"].is_unique
#assert "incucyte timestamp" in df.columns


assert "compounds" in dfs_per_sheetname
df = dfs_per_sheetname["compounds"]
assert "compound ID"  in df.columns
assert "SMILES"         in df.columns

df2 = df[df.duplicated(subset=["SMILES"], keep=False)]
if len(df2) > 0:
    print("Sheet 'compounds': The following groups of entries have the same SMILES but different compound IDs:")
    for g, s in df2.groupby("SMILES"):
        print(f"{g} : ")
        print(s)
    print("---")

df2 = df[df.duplicated(subset=["compound ID"], keep=False)]
if len(df2) > 0:
    print("Sheet 'compounds': The following groups of entries have the same compound ID but different SMILES:")
    for g, s in df2.groupby("compound ID"):
        print(f"{g} : ")
        print(s)
    print("---")

assert df["compound ID"].is_unique
assert df["SMILES"].is_unique

assert not df["SMILES"].str.contains("\n").any()




assert "compound batches" in dfs_per_sheetname
df = dfs_per_sheetname["compound batches"]
assert "compound batch ID"  in df.columns
assert "compound ID"      in df.columns

df2 = df[df.duplicated(subset=["compound batch ID"], keep=False)]
if len(df2) > 0:
    print("Sheet 'compound batches': The following groups of entries have the same compound batch ID:")
    for g, s in df2.groupby("compound batch ID"):
        print(f"{g} : ")
        print(s)
    print("---")

assert df["compound batch ID"].is_unique




mapping_tables_to_check = list( [s for s in dfs_per_sheetname if "compound map" in s] )
for mapping_table_name in mapping_tables_to_check:
    assert mapping_table_name in dfs_per_sheetname
    df = dfs_per_sheetname[mapping_table_name]
    assert "well ID"  in df.columns
    assert "well name"      in df.columns
    assert "compound batch ID"      in df.columns
    assert "concentration uM" in df.columns
    assert "experimental type" in df.columns


## complex tests follow...
acceptable_experimental_types = ["chemogenomic candidate", "unrelated to this experiment", "blank", "control", "cells only"]

for mapping_table_name in mapping_tables_to_check:
    df = dfs_per_sheetname[mapping_table_name]

    ## check that all rows contain one of the allowed values above
    assert df["experimental type"].isin(acceptable_experimental_types).all()


    # concentration should be only nan if experimental type is one of the below
    cond1 = df["experimental type"] == "blank"
    cond2 = df["concentration uM"].isna()
    cond3 = df["experimental type"] == "cells only"
    cond3b = df["experimental type"] == "unrelated to this experiment"
    assert df[cond1].equals(df[(cond1)  & (cond2)])
    assert df[cond3].equals(df[(cond3)  & (cond2)])
    assert df[cond3b].equals(df[(cond3b) & (cond2)])
    assert df[cond2].equals(df[(cond1) | (cond3) | (cond3b)])

    # concentration should be >0 if experimental type is different than the ones above
    df_out = df[~((cond1)|(cond3)|(cond3b))].query("not `concentration uM` > 0")
    if len( df_out ) > 0:
        print(f"Concentrations in table '{mapping_table_name}' are not in the expected range:")
        print(df_out)
        print("---")

    # compound batch should be only nan if experimental type is one of the above
    cond4 = df["compound batch ID"].isna()
    assert df[cond1].equals(df[(cond4) & (cond1)])
    assert df[cond3].equals(df[(cond4) & (cond3)])
    assert df[cond3b].equals(df[(cond4) & (cond3b)])
    assert df[cond4].equals(df[(cond1) | (cond3) | (cond3b)])


## ID reference tests
foo = dfs_per_sheetname["experiments"]["experiment ID"]
bar = dfs_per_sheetname["imaging campaigns"]["experiment ID"]
assert foo.isin(bar.values).all()
assert bar.isin(foo.values).all()


foo = dfs_per_sheetname["compound batches"]["compound ID"]
bar = dfs_per_sheetname["compounds"]["compound ID"]
bar_foo = set(bar) - set(foo)
if len(bar_foo) > 0:
    print("INFO: There are compound IDs in table 'compounds', which are not referenced in table 'compound batches':")
    print(bar_foo)
    print("---")
foo_bar = set(foo) - set(bar)
if len(foo_bar) > 0:
    print("There are compound IDs in table 'compound batches', which cannot be resolved from table 'compounds':")
    print(foo_bar)
    print("---")
assert foo.isin(bar.values).all()
assert bar.isin(foo.values).all()



for mapping_table_name in mapping_tables_to_check:
    foo = dfs_per_sheetname["compound batches"]["compound batch ID"].unique()
    bar = dfs_per_sheetname[mapping_table_name]
    bar = bar[ bar["experimental type"] != "cells only" ]
    bar = bar[ bar["experimental type"] != "blank"]
    bar = bar[ bar["experimental type"] != "unrelated to this experiment"]
    bar = bar["compound batch ID"].unique()

    bar_foo = set(bar) - set(foo)
    if len(bar_foo) > 0:
        print(f"There are compound batches in table '{mapping_table_name}', which cannot be resolved from table 'compound batches':")
        print(bar_foo)
        print("---")

print("Done.")

## BLOCK to replace dummy values in the whole excel file
if True:

    did_i_change_anything = False

    mapping = {
        # "old" : "new",
        "dummy1" : "dummy1",
        "dummy2" : "EUB0001080a",
        "dummy3" : "DP000007a",
        "dummy4" : "EUB0001108a",

        "EUB0000500a" : "EUB0000871a",
        "EUB0000528a" : "EUB0000841a",
        "EUB0000543aCl" : "EUB0000213bCl",
        "EUB0000550aCl" : "EUB0000196bCl",
        "EUB0000657aPO4" : "EUB0000140bPO4",
        "EUB0000667aCit" : "EUB0000286bCit",
        "EUB0000675aCl" : "EUB0000130bCl",
        "EUB0000092a" : "EUB0000092b"
    }

    import openpyxl
    wb = openpyxl.load_workbook(FILENAME)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        dimension = ws.calculate_dimension()
        for row in ws[dimension]:
            for cell in row:
                if cell.value in mapping:
                    print(f"Changing cell {cell} from value {cell.value} to {mapping[cell.value]}")
                    cell.value = mapping[cell.value]
                    did_i_change_anything = True

    if did_i_change_anything:
        wb.save(FILENAME + ".changed.xlsx")
## ... end of BLOCK.


## BLOCK to check the whole excel file for trailing spaces in the fields
if True:
    import openpyxl
    wb = openpyxl.load_workbook(FILENAME)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        dimension = ws.calculate_dimension()
        for row in ws[dimension]:
            for cell in row:
                if type(cell.value) == str and cell.value.strip() != cell.value:
                    print(f"Sheet '{sheetname}', cell {cell.coordinate} contains undesired whitespace: '{cell.value}'")
## ... end of BLOCK.

## BLOCK to condense a list of superfluous entries in table 'compounds' vs correct table 'compound batches'
if False:
    foo = dfs_per_sheetname["compound batches"]["compound ID"]
    bar = dfs_per_sheetname["compounds"]["compound ID"]
    bar_foo = set(bar) - set(foo)
    dfs_per_sheetname["compounds"][~bar.isin(bar_foo)].to_excel("2022-02-03-new-compounds-sheet.xlsx")
## ... end of BLOCK.

## BLOCK to check for expected pattern in compound concentrations in one plate...
if False:
    for mapping_table_name in mapping_tables_to_check:
        foo = dfs_per_sheetname[mapping_table_name]
        foo = foo[foo["experimental type"]=="chemogenomic candidate"]
        print(mapping_table_name)
        print("total len:",len(foo))
        counter=0
        for groupname, series in foo.groupby("eubopen ID"):
            if len(series)!=2:
                if len(series)==1:
                    if series["concentration uM"].item() == 10.0:
                        counter+=1
                        continue
                print("potential ERROR:")
                print(series)
            else:
                if sorted(series["concentration uM"].values) == [1.0, 10.0]:
                    counter+=2
                else:
                    print("potential ERROR:")
                    print(series)
        print("rather unsuspicious:", counter)
## ... end of BLOCK.


### BLOCK to check for consistency in data and produce condensed output, if EUbOPEN, SGC IDs, and compound names are given in the compound maps ...
if False:
    collect_mappings_between_sgc_and_eubopen_id = {}
    collect_mappings_between_compound_names_and_eubopen_id = {}
    for mapping_table_name in mapping_tables_to_check:
        spam = dfs_per_sheetname[mapping_table_name][["SGC ID", "eubopen ID"]].dropna().drop_duplicates()
        spam = dfs_per_sheetname[mapping_table_name][["SGC ID", "eubopen ID"]].drop_duplicates()
        same_sgc_different_eubopen = spam[spam.duplicated(subset="SGC ID", keep=False)]
        same_eubopen_different_sgc = spam[spam.duplicated(subset="eubopen ID", keep=False)]
        if len(same_eubopen_different_sgc)>0:
            print(f"There are compound batches in table '{mapping_table_name}', which have different SGC IDs, but the same EUbOPEN ID:")
            print(same_eubopen_different_sgc)
            print("---")
        if len(same_sgc_different_eubopen)>0:
            print(f"There are compound batches in table '{mapping_table_name}', which have the same SGC ID, but different EUbOPEN IDs:")
            print(same_sgc_different_eubopen)
            print("---")
        #assert len(same_sgc_different_eubopen) == 0
        #assert len(same_eubopen_different_sgc) == 0

        for sgc_id, s in spam.groupby("SGC ID"):
            if sgc_id in collect_mappings_between_sgc_and_eubopen_id:
                value = s["eubopen ID"].item()
                if value != collect_mappings_between_sgc_and_eubopen_id[sgc_id] and not (pandas.isna(value) and pandas.isna(collect_mappings_between_sgc_and_eubopen_id[sgc_id])):
                    print(f"ERROR for {sgc_id}: {repr(value)} != {repr(collect_mappings_between_sgc_and_eubopen_id[sgc_id])}")
            else:
                collect_mappings_between_sgc_and_eubopen_id.update( {sgc_id: s["eubopen ID"].item()} )

        spam2 = dfs_per_sheetname[mapping_table_name][["compound ID", "eubopen ID"]].drop_duplicates()
        for compound_name, s in spam2.groupby("compound ID"):
            if pandas.isna(compound_name) or len(s)>1:
                print(f"compound name is nan: {s}")
            if compound_name in collect_mappings_between_compound_names_and_eubopen_id:
                value = s["eubopen ID"].item()
                if value != collect_mappings_between_compound_names_and_eubopen_id[compound_name] and not (pandas.isna(value) and pandas.isna(collect_mappings_between_compound_names_and_eubopen_id[compound_name])):
                    print(f"ERROR for {compound_name}: {repr(value)} != {repr(collect_mappings_between_compound_names_and_eubopen_id[compound_name])}")
            else:
                collect_mappings_between_compound_names_and_eubopen_id.update( {compound_name: s["eubopen ID"].item()} )


    print(collect_mappings_between_sgc_and_eubopen_id)
    print(collect_mappings_between_compound_names_and_eubopen_id)

    df1 = pandas.DataFrame.from_dict(collect_mappings_between_compound_names_and_eubopen_id, orient="index", columns=["eubopen ID"])
    df1["compound ID"] = df1.index
    #print(df1[df1.duplicated("eubopen ID", keep=False)])

    df2 = pandas.DataFrame.from_dict(collect_mappings_between_sgc_and_eubopen_id, orient="index", columns=["eubopen ID"])
    df2["SGC ID"] = df2.index
    #print(df2[df2.duplicated("eubopen ID", keep=False)])

    df3 = df1.dropna().merge(df2.dropna(), how="left", on="eubopen ID", validate="1:1").sort_values("eubopen ID")
    df3.to_csv("2022-02-02-df3.csv")


    #print(df3[~df3["compound ID"].isin(dfs_per_sheetname["compounds"]["compound ID"].values)])
### ... end of BLOCK.

### BLOCK to check compound name and SMILES correlations within database for unexpected effects ...
if False:
    df = dfs_per_sheetname["compounds old"]
    df.drop_duplicates(subset=["compound ID"], keep="first")[["compound ID", "SMILES"]].sort_values("compound ID").to_excel("2022-02-01-compounds-out.xlsx")
    df[["SGC Global Compound ID (Batch)", "compound ID"]].sort_values("SGC Global Compound ID (Batch)").to_excel("2022-02-01-compound-batches-out.xlsx")
    print("---")
    print("these entries have different SMILES for the same compound name:")
    for groupname, s in df[df.duplicated(subset=["compound ID"], keep=False)].groupby("compound ID"):
        different= s.drop_duplicates(subset=["compound ID", "SMILES"], keep=False)
        if len(different)>0:
            print(different)
    print("---")
    print("these entries have different names but the same SMILES:")
    df2 = df[["compound ID", "SMILES"]].dropna(axis=0,subset=["SMILES"])
    df3 = df2[df2.duplicated(subset=["SMILES"], keep=False)]
    for i,s in df3.groupby("SMILES"):
        different= s.drop_duplicates(subset=["compound ID", "SMILES"], keep=False)
        if len(different)>0:
            print(different)
    print("---")
    print("these entries have no SMILES:")
    df2 = df[df["SMILES"].isna()]
    print(df2)
    print("---")
    print("these entries have no batch number:")
    print(df[df["SGC Global Compound ID (Batch)"].isna()])
    print("---")
### ... end of BLOCK.

